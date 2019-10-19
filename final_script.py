from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep
from parsel import Selector
import openpyxl

driver=webdriver.Chrome('C:/Users/gunjan/Desktop/Web_Scraping/chromedriver')
driver.get('https://www.linkedin.com/login?fromSignIn=true&trk=guest_homepage-basic_nav-header-signin')
sleep(0.5)
uname=driver.find_element_by_id('username')
uname.send_keys('9955867077')
sleep(0.5)
passwd=driver.find_element_by_id('password')
passwd.send_keys('Aadi@123')
sleep(0.5)
sign_in_btn=driver.find_element_by_xpath('//*[@type="submit"]')
sign_in_btn.click()
sleep(5)


driver.get('https://www.linkedin.com/search/results/people/?keywords=CTO&origin=SWITCH_SEARCH_VERTICAL')
sleep(5)
driver.refresh()
#driver.get('https://www.linkedin.com/search/results/people/?keywords=SDE&origin=SWITCH_SEARCH_VERTICAL')
sel=Selector(text=driver.page_source)
#st=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "actor-name", " " ))]/text()').extract()
st=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "actor-name", " " ))]/text()').extract()
#st=sel.css('.actor-name').extract()
jt=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "t-black", " " ))]//span/text()').extract()
pt=sel.xpath('//*[(@id = "ember62")]//*[contains(concat( " ", @class, " " ), concat( " ", "t-black--light", " " ))]//span/text()').extract()

path="C:/Users/gunjan/Desktop/Web_Scraping/data.xlsx"

workbook=openpyxl.load_workbook(path)
sheet=workbook.active


for r in range(1,len(st)):
    #for c in range(1,4):
    sheet.cell(row=r,column=1).value=st[r-1]
        #i=i+1

for r in range(1,len(st)):
    #for c in range(1,4):
    sheet.cell(row=r,column=2).value=jt[r-1]
        #i=i+1

for r in range(1,len(st)):
    #for c in range(1,4):
    sheet.cell(row=r,column=3).value=pt[r-1]
        #i=i+1


workbook.save(path)


#np=driver.find_element_by_xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "artdeco-button__text", " " ))]')
#np.click()
'''
driver.get('https://www.linkedin.com/search/results/people/?keywords=CTO&origin=SWITCH_SEARCH_VERTICAL&page=2')

sen=Selector(text=driver.page_source)
sn=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "actor-name", " " ))]/text()').extract()
jn=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "t-black", " " ))]//span/text()').extract()

#len(sn)
#len(jn)

path="C:/Users/gunjan/Desktop/Web_Scraping/data2.xlsx"


workbook=openpyxl.load_workbook(path)
sheet=workbook.active

#i=0
for r in range(1,len(sn)):
    #for c in range(1,4):
    sheet.cell(row=r,column=1).value=sn[r-1]
    #i=i+1
        #i=i+1
#i=0
for r in range(1,len(jn)):
    #for c in range(1,4):
    sheet.cell(row=r,column=2).value=jn[r-1]
    #i=i+1
        #i=i+1

workbook.save(path)
'''
